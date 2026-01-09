"""
Tool to read and analyze CSV and Excel files using pandas.
Creates a JSON packet with column statistics.
"""

import json
import re
import pandas as pd
import numpy as np
from openpyxl import load_workbook


def detect_pii_pattern(series: pd.Series) -> tuple[bool, str, str]:
    """
    Detect if column contains PII and return pattern type.
    
    Args:
        series: Pandas series to analyze
        
    Returns:
        Tuple of (is_pii, pattern_description, pattern_type)
    """
    sample = series.dropna().astype(str).head(50)
    if len(sample) == 0:
        return False, "No data", "none"
    
    email_pattern = r'^[a-zA-Z0-9._%+-]+@[a-zA-Z0-9.-]+\.[a-zA-Z]{2,}$'
    phone_pattern = r'(?:\d{3}[-.\s]?\d{3}[-.\s]?\d{4}|\(\d{3}\)\s?\d{3}[-.\s]?\d{4})'
    zip_us_pattern = r'^\d{5}(?:-\d{4})?$'
    postcode_uk_pattern = r'^[A-Z]{1,2}\d{1,2}[A-Z]?\s?\d[A-Z]{2}$'
    
    email_matches = sample.str.match(email_pattern, na=False).sum()
    phone_matches = sample.str.contains(phone_pattern, na=False, regex=True).sum()
    zip_matches = sample.str.match(zip_us_pattern, na=False).sum()
    postcode_matches = sample.str.match(postcode_uk_pattern, na=False, case=False).sum()
    
    threshold = len(sample) * 0.7
    
    if email_matches > threshold:
        return True, "Email addresses", "email"
    elif phone_matches > threshold:
        return True, "Phone numbers", "phone"
    elif zip_matches > threshold:
        return True, "US ZIP codes", "zip"
    elif postcode_matches > threshold:
        return True, "UK Postcodes", "postcode"
    
    address_keywords = sample.str.lower().str.contains(
        r'\b(?:street|st|avenue|ave|road|rd|drive|dr|lane|ln|boulevard|blvd|way|court|ct)\b',
        na=False, regex=True
    ).sum()
    if address_keywords > threshold:
        return True, "Street addresses", "address"
    
    return False, "", "none"


def generate_fake_samples(pattern_type: str, count: int = 3) -> list[str]:
    """Generate fake samples for PII data."""
    if pattern_type == "email":
        return [f"user{i}@example.com" for i in range(1, count + 1)]
    elif pattern_type == "phone":
        return [f"555-01{i:02d}" for i in range(count)]
    elif pattern_type == "zip":
        return ["12345", "67890", "11111"][:count]
    elif pattern_type == "postcode":
        return ["SW1A 1AA", "EC1A 1BB", "W1A 0AX"][:count]
    elif pattern_type == "address":
        return ["123 Main St", "456 Oak Ave", "789 Pine Rd"][:count]
    else:
        return ["[REDACTED - PII]"] * count


def normalize_data_type(dtype) -> str:
    """Normalize pandas dtype to simplified type for LLM."""
    dtype_str = str(dtype).lower()
    if 'int' in dtype_str:
        return "integer"
    elif 'float' in dtype_str:
        return "float"
    elif 'datetime' in dtype_str or 'date' in dtype_str:
        return "date"
    else:
        return "string"


def get_cardinality_level(percent_unique: float, unique_count: int) -> str:
    """Determine cardinality level based on uniqueness."""
    if unique_count == 1:
        return "single"
    elif percent_unique >= 99:
        return "unique"
    elif percent_unique >= 50:
        return "high"
    elif percent_unique >= 10:
        return "medium"
    else:
        return "low"


def has_strong_pattern(series: pd.Series) -> bool:
    """Check if series has a strong structural pattern (like ZIP codes, phone numbers)."""
    sample = series.dropna().astype(str).head(100)
    if len(sample) == 0:
        return False
    
    zip_us_pattern = r'^\d{5}(?:-\d{4})?$'
    phone_pattern = r'(?:\d{3}[-.\s]?\d{3}[-.\s]?\d{4}|\(\d{3}\)\s?\d{3}[-.\s]?\d{4})'
    postcode_uk_pattern = r'^[A-Z]{1,2}\d{1,2}[A-Z]?\s?\d[A-Z]{2}$'
    
    zip_matches = sample.str.match(zip_us_pattern, na=False).sum()
    phone_matches = sample.str.contains(phone_pattern, na=False, regex=True).sum()
    postcode_matches = sample.str.match(postcode_uk_pattern, na=False, case=False).sum()
    
    threshold = len(sample) * 0.7
    
    return (zip_matches > threshold or phone_matches > threshold or postcode_matches > threshold)


def determine_likely_role(is_date: bool, is_numeric: bool, data_type: str, percent_populated: float, 
                          percent_unique: float, cardinality: str, is_empty_numeric: bool, has_pattern: bool) -> str:
    """Determine the likely role of the column."""
    if percent_populated == 100 and percent_unique == 100 and not is_date:
        return "primary_key"
    elif is_date:
        return "date"
    elif not is_date and data_type != "float" and cardinality in ["high", "unique"] and percent_unique > 50:
        if has_pattern:
            return "categorical"
        return "foreign_key"
    elif is_numeric and not is_empty_numeric:
        return "numeric_measure"
    elif cardinality == "low" and not is_numeric:
        return "categorical"
    elif is_numeric:
        return "numeric_measure"
    else:
        return "text"


def check_sentinel_dates(series: pd.Series) -> str:
    """Check for sentinel date values."""
    if not pd.api.types.is_datetime64_any_dtype(series):
        return "none"
    
    sample = series.dropna().head(100)
    if len(sample) == 0:
        return "none"
    
    has_low = False
    has_high = False
    
    for date in sample:
        year = pd.to_datetime(date).year
        if year <= 1900 or year == 1970:
            has_low = True
        elif year >= 9999:
            has_high = True
    
    if has_low and has_high:
        return "both"
    elif has_low:
        return "low"
    elif has_high:
        return "high"
    else:
        return "none"


def get_numeric_range(series: pd.Series) -> dict:
    """Get min and max for numeric columns."""
    clean_data = series.dropna()
    if len(clean_data) == 0:
        return {}
    
    min_val = float(clean_data.min())
    max_val = float(clean_data.max())
    
    return {
        "min": min_val,
        "max": max_val
    }


def get_date_range(series: pd.Series) -> dict:
    """Get min and max for date columns, excluding sentinel values."""
    clean_data = series.dropna()
    if len(clean_data) == 0:
        return {}
    
    if pd.api.types.is_datetime64_any_dtype(series):
        date_series = clean_data
    else:
        date_series = pd.to_datetime(clean_data, errors='coerce')
    
    filtered = date_series[(date_series.dt.year > 1900) & (date_series.dt.year < 9999)]
    
    if len(filtered) == 0:
        return {}
    
    min_date = filtered.min()
    max_date = filtered.max()
    
    return {
        "min": str(min_date.date() if hasattr(min_date, 'date') else min_date),
        "max": str(max_date.date() if hasattr(max_date, 'date') else max_date)
    }


def check_all_times_zero(series: pd.Series) -> bool:
    """Check if all time components are 00:00:00 in datetime column."""
    if not pd.api.types.is_datetime64_any_dtype(series):
        return False
    
    clean_data = series.dropna()
    if len(clean_data) == 0:
        return False
    
    date_series = pd.to_datetime(clean_data)
    return ((date_series.dt.hour == 0).all() and 
            (date_series.dt.minute == 0).all() and 
            (date_series.dt.second == 0).all())


def infer_pattern(series: pd.Series, is_numeric: bool) -> str:
    """Infer pattern from non-PII column data."""
    if is_numeric:
        sample = series.dropna().head(100)
        if len(sample) == 0:
            return "No data"
        
        sample_str = sample.astype(str)
        has_decimal = sample_str.str.contains(r'\.', regex=True)
        
        if has_decimal.any():
            decimal_parts = sample_str.str.extract(r'\.(\d+)')
            decimal_lengths = decimal_parts[0].dropna().str.len()
            
            if len(decimal_lengths) > 0:
                unique_lengths = decimal_lengths.unique()
                if len(unique_lengths) == 1:
                    return f"Decimal precision: {int(unique_lengths[0])} places"
                else:
                    min_dp = int(decimal_lengths.min())
                    max_dp = int(decimal_lengths.max())
                    return f"Decimal precision: {min_dp}-{max_dp} places"
        
        return "Integer values"
    
    sample = series.dropna().astype(str).head(100)
    if len(sample) == 0:
        return "No data"
    
    def create_pattern(s: str) -> str:
        pattern = s
        pattern = re.sub(r'\d', '#', pattern)
        pattern = re.sub(r'[a-zA-Z]', 'X', pattern)
        return pattern
    
    patterns = sample.apply(create_pattern)
    pattern_counts = patterns.value_counts()
    
    if len(pattern_counts) == 1:
        return f"Consistent: {pattern_counts.index[0]}"
    elif len(pattern_counts) <= 3:
        top_patterns = pattern_counts.head(3).index.tolist()
        return f"Common patterns: {', '.join(top_patterns)}"
    else:
        most_common = pattern_counts.head(1).index[0]
        coverage = (pattern_counts.head(1).values[0] / len(sample)) * 100
        if coverage > 70:
            return f"Dominant: {most_common} ({coverage:.0f}%)"
        return "No clear pattern"


def detect_actual_type(series: pd.Series) -> tuple[str, bool]:
    """
    Detect actual data type by analyzing content, overriding pandas if needed.
    Returns (actual_type, is_mixed)
    """
    if series.dtype != 'object':
        return normalize_data_type(series.dtype), False
    
    sample = series.dropna().astype(str)
    sample = sample[sample.str.strip() != '']
    
    if len(sample) == 0:
        return "string", False
    
    date_pattern = r'^\d{4}-\d{2}-\d{2}(?:\s\d{2}:\d{2}:\d{2})?$'
    date_matches = sample.str.match(date_pattern).sum()
    if date_matches / len(sample) > 0.8:
        return "date", False
    
    def is_numeric_string(s: str) -> bool:
        s = s.strip()
        try:
            float(s)
            return True
        except ValueError:
            if re.match(r'^-?\d+\.?\d*[eE][+-]?\d+$', s):
                return True
            return False
    
    numeric_mask = sample.apply(is_numeric_string)
    numeric_count = numeric_mask.sum()
    numeric_ratio = numeric_count / len(sample)
    
    if numeric_ratio == 1.0:
        if sample.str.contains(r'[eE]').any():
            return "float", False
        elif sample.str.contains(r'\.').any():
            return "float", False
        else:
            return "integer", False
    elif numeric_ratio > 0.5:
        return "string", True
    
    return "string", False


def check_mixed_type(series: pd.Series) -> bool:
    """Check if column contains mixed data types."""
    _, is_mixed = detect_actual_type(series)
    return is_mixed


def read_and_analyze(file_path: str, sheet_name: str | None = None) -> dict:
    """
    Read a CSV or Excel file and create column statistics.

    Args:
        file_path: Path to the CSV or Excel file
        sheet_name: Sheet name for Excel files (optional, uses first sheet if not provided)

    Returns:
        Dictionary with column statistics including data_type, percent_populated, percent_unique, and unique_count
    """
    if file_path.endswith(".csv"):
        df = pd.read_csv(file_path)
        print(f"\n=== Analyzing CSV file: {file_path} ===\n")
    elif file_path.endswith((".xlsx", ".xls")):
        if sheet_name is None:
            try:
                wb = load_workbook(file_path, read_only=True)
                sheet_name = wb.sheetnames[0]
                wb.close()
                print(f"\n=== Using first sheet: {sheet_name} ===\n")
            except Exception:
                sheet_name = None  # Will default to first sheet
                print(f"\n=== Using first sheet (index 0) ===\n")

        df = pd.read_excel(file_path, sheet_name=sheet_name)
        print(f"\n=== Analyzing Excel file: {file_path} ===\n")
    else:
        raise ValueError(f"Unsupported file type: {file_path}")

    for col in df.columns:
        if df[col].dtype == 'object':
            df[col] = df[col].apply(lambda x: x.strip() if isinstance(x, str) else x)
    
    df = df.replace('', pd.NA)
    
    total_rows = len(df)
    column_stats = {}

    for column in df.columns:
        non_null_count = df[column].notna().sum()
        unique_count = df[column].nunique()
        
        is_date = pd.api.types.is_datetime64_any_dtype(df[column])
        is_numeric = pd.api.types.is_numeric_dtype(df[column])
        percent_populated = (non_null_count / total_rows) * 100 if total_rows > 0 else 0
        percent_unique = (unique_count / total_rows) * 100 if total_rows > 0 else 0
        
        data_type, mixed_type = detect_actual_type(df[column])
        
        if data_type == "date":
            is_date = True
        if data_type in ["integer", "float"]:
            is_numeric = True
        
        is_empty_numeric = is_numeric and percent_populated < 1.0
        has_pattern = has_strong_pattern(df[column]) if data_type == "string" else False
        
        cardinality = get_cardinality_level(percent_unique, unique_count)
        likely_role = determine_likely_role(is_date, is_numeric, data_type, percent_populated, percent_unique, cardinality, is_empty_numeric, has_pattern)
        
        is_pii, pii_description, pii_type = detect_pii_pattern(df[column])
        
        if is_pii:
            pattern = pii_description
            random_samples = generate_fake_samples(pii_type, min(3, len(df[column].dropna())))
        else:
            pattern = infer_pattern(df[column], is_numeric)
            sample_data = df[column].dropna()
            non_blank_data = sample_data[sample_data.astype(str).str.strip() != '']
            
            pseudo_nulls = ['nan', 'NaN', 'null']
            non_blank_data = non_blank_data[~non_blank_data.astype(str).isin(pseudo_nulls)]
            
            if len(non_blank_data) == 0:
                random_samples = []
            else:
                value_counts = non_blank_data.value_counts()
                total_values = len(non_blank_data)
                
                random_samples = []
                
                if len(value_counts) > 1:
                    top_value = value_counts.index[0]
                    top_count = value_counts.iloc[0]
                    top_percentage = (top_count / total_values) * 100
                    
                    if top_percentage > 50:
                        random_samples.append(str(top_value))
                        
                        other_values = non_blank_data[non_blank_data != top_value]
                        if len(other_values) > 0:
                            sample_count = min(2, len(other_values.unique()))
                            if sample_count > 0:
                                other_samples = other_values.drop_duplicates().sample(sample_count, random_state=42)
                                random_samples.extend(other_samples.astype(str).tolist())
                    else:
                        sample_count = min(3, len(non_blank_data.unique()))
                        random_samples = non_blank_data.drop_duplicates().sample(sample_count, random_state=42).astype(str).tolist()
                else:
                    random_samples = [str(non_blank_data.iloc[0])]
        
        if is_date and check_all_times_zero(df[column]):
            actual_data_type = "date"
        elif is_date:
            actual_data_type = "datetime"
        else:
            actual_data_type = data_type
        
        col_stats = {
            "data_type": actual_data_type,
            "likely_role": likely_role,
            "percent_populated": float(round(percent_populated, 2)),
            "percent_unique": float(round(percent_unique, 2)),
            "unique_count": int(unique_count),
            "cardinality_level": cardinality,
            "pattern": str(pattern),
            "samples": [str(s) for s in random_samples],
            "mixed_type": bool(mixed_type)
        }
        
        if is_numeric:
            zero_count = (df[column] == 0).sum()
            col_stats["zero_count"] = int(zero_count)
            col_stats["zero_percent"] = float(round((zero_count / total_rows) * 100, 2)) if total_rows > 0 else 0.0
            
            num_range = get_numeric_range(df[column])
            if num_range:
                col_stats["min"] = num_range["min"]
                col_stats["max"] = num_range["max"]
        
        if is_date:
            sentinel_type = check_sentinel_dates(df[column])
            if sentinel_type != "none":
                col_stats["sentinel_date_type"] = sentinel_type
            
            date_range = get_date_range(df[column])
            if date_range:
                col_stats["min"] = date_range["min"]
                col_stats["max"] = date_range["max"]
        
        column_stats[column] = col_stats

    result = {
        "file_path": file_path,
        "total_rows": total_rows,
        "total_columns": len(df.columns),
        "columns": column_stats
    }

    print(json.dumps(result, indent=2))
    return result


if __name__ == "__main__":
    # Hardcoded file paths for manual testing
    file_path = r"C:\Users\simon2282\OneDrive - Willis Towers Watson\Documents\Dev\projects\column-matcher\data\reference_small.xlsx"  # Change this to your file path
    sheet_name = None  # Change this if you want a specific sheet

    read_and_analyze(file_path, sheet_name)
